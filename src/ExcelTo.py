#!/usr/bin/env python
# encoding:utf-8
from distutils.fancy_getopt import FancyGetopt
from math import fabs
import os
from openpyxl import load_workbook 
import time
import getopt
import sys

from CodeCreator.TTSCreator import TTSCreator
from CodeCreator.TFormulaCreator import TFormulaCreator
from CodeCreator.TLanguageCreator import TLanguageCreator


def getKeyListFromWS(ws):
    """
    从配置表的第一行数据中获取主键信息，支持单一主键，支持联合主键
    """    
    keyList = []
    keyConfig = ws[1][0].value                  # 提取主键数据
    keyConfig = keyConfig.replace("key:", "")   # 删除字符串开头key:
    keyList = keyConfig.split(",")              # 把字符串切割放到数组中
    keyList = list(map(int, keyList))           # 把数组元素变成数值类型
    keyList = sorted(keyList)                   # 主键排序
    return keyList


def getFieldListFromWS(ws, creator):
    fieldList = []
    maxCol = 0
    # 前3行为表说明 主键 字段 注释
    for col in ws[4]:
        if (col.value != None):
            aField = creator.creatorField(col.value, ws[3][maxCol].value, ws[2][maxCol].value)
            fieldList.append(aField)
            maxCol += 1
        else:
            break
    return fieldList


def createCode(path, exportDir, creator):
    """
    生成代码
    :param path: xlsx文件路径
    :param exportDir: 生成文件路径
    :param creator: 生成器对象
    """
    ws = load_workbook(filename=path, read_only=True)["Sheet1"]
    fileName = os.path.splitext(os.path.basename(path))[0]    
    tsFilePath = os.path.join(exportDir, fileName + creator.getFileSuffix())

    maxRows = ws.max_row  # 行
    aFile = open(tsFilePath, "w", encoding="utf-8")
    keyList = getKeyListFromWS(ws)
    fieldList = getFieldListFromWS(ws, creator)

    creator.init(fileName, fieldList, keyList)
    aFile.write(creator.ceatorAllCode(ws, len(fieldList), maxRows))


def excelPathGenerator(excelDir):
    """
    生成器 - 用于遍历文件目录找出后缀名为xlsx或xls的文件
    :param excelDir: 需要便利的目录
    """
    for parent, _, filenames in os.walk(excelDir):
        for filename in filenames:
            if filename.startswith("~") or (not filename.endswith(".xlsx")
                                            and not filename.endswith(".xls")):
                continue
            path = os.path.join(parent, filename)
            yield path


def dealDataExcel(dataExcelDir, outPath, outTypeList):
    """
    处理数据表excel
    """
    if len(outTypeList) < 1:
        print("Please specify the format")
        return

    print("开始生成配置表")
    f = excelPathGenerator(dataExcelDir)
    while True:
        try:
            path = next(f)            
            t1 = int(round(time.time() * 1000))            

            if "ts" in outTypeList:
                if not os.path.exists(outPath):
                    os.makedirs(outPath)
                createCode(path, outPath, TTSCreator())            
            
            t2 = int(round(time.time() * 1000))
            print("-> " + path + " time: " + str(t2 - t1) + "ms")
        except StopIteration:
            break


def dealFormulaExcel(path, outPath, outTypeList):
    """
    处理公式excel
    """
    if path:
        ws = load_workbook(filename=path, read_only=True)["Sheet1"]
        if "ts" in outTypeList:
            outPath = os.path.join(outPath, "formula.ts")
            creator = TFormulaCreator()
            with open(outPath, 'w') as load_f:
                load_f.write(creator.creatorTS(ws, 4, ws.max_row))        


def dealLanguageExcel(path, outDir):
    """
    处理多语言表
    """
    ws = load_workbook(filename=path, read_only=True)["Sheet1"]
    creator = TLanguageCreator()
    saveToFile(os.path.join(outDir, "zh_cn.json"),
               creator.creatorsZHLanguage(ws, 4, ws.max_row))
    saveToFile(os.path.join(outDir, "en.json"),
               creator.creatorsENLanguage(ws, 4, ws.max_row))
    

def saveToFile(outPath, data):
    with open(outPath, 'w', encoding="utf-8") as load_f:
        load_f.write(data)


if __name__ == "__main__":
    try:
        opts, args = getopt.getopt(sys.argv[1:], "ht:f:o:l:", ["ts", "language"])            
    except getopt.GetoptError:
        sys.exit(2)

    dataExcelDir = ""
    formulaExcelPath = ""
    languageExcelPath = ""
    outDir = ""
    sType = []
    for opt, arg in opts:
        if opt == '-h':
            print("-o 导出路径")
            print("-t 数据表excel文件夹路径")
            print("-f 公式excel文件路径,excel文件")
            print("-l language文件路径,excel文件")
            print("--ts 导出ts")                
            print("--language 导出多语言配置")
            sys.exit()
        elif opt == "-t":
            dataExcelDir = arg
        elif opt == "-f":
            formulaExcelPath = arg
        elif opt == "-l":
            languageExcelPath = arg
        elif opt == "-o":
            outDir = arg
        elif opt == "--ts":
            sType = "ts"            
        elif opt == "--language":
            sType = "language"

    # 数据表导出
    if dataExcelDir != "" and outDir != "":
        if not os.path.exists(dataExcelDir):
            print("excel dir not exist", dataExcelDir)
            sys.exit(2)
        print("exce to txt:")
        dealDataExcel(dataExcelDir, outDir, [sType])

    if formulaExcelPath != "" and outDir != "":
        if not os.path.exists(formulaExcelPath):
            print("formulaExcel not exist", formulaExcelPath)
            sys.exit(2)
        print("exce to Formula:")
        dealFormulaExcel(formulaExcelPath, outDir, [sType])

    if languageExcelPath != "" and outDir != "":
        if not os.path.exists(languageExcelPath):
            print("formulaExcel not exist", languageExcelPath)
            sys.exit(2)
        print("exce to language:")
        dealLanguageExcel(languageExcelPath, outDir)
    
